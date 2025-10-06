# =============================================================================
# PERFORMANCE_LOGGER.PY - MODULE THỐNG KÊ VÀ XUẤT KẾT QUẢ RA FILE EXCEL
# =============================================================================
# Module này quản lý thống kê hiệu suất của các thuật toán AI và xuất ra file Excel
# Hỗ trợ so sánh hiệu suất giữa BFS, DFS, UCS, A*, Greedy, v.v.

import os
import time
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class PerformanceLogger:
    """
    Class quản lý thống kê hiệu suất của các thuật toán AI
    - Ghi nhận dữ liệu từ mỗi ván chơi
    - Xuất kết quả ra file Excel với format đẹp
    - Hỗ trợ append dữ liệu mới vào file hiện có
    """
    
    def __init__(self, excel_file="algorithm_comparison.xlsx"):
        """
        Khởi tạo PerformanceLogger
        Args:
            excel_file: Tên file Excel để lưu kết quả
        """
        self.excel_file = excel_file
        self.data_records = []  # Danh sách các bản ghi thống kê
        self.current_game_data = {}  # Dữ liệu game hiện tại
        
        # Khởi tạo headers cho Excel
        self.headers = [
            "Algorithm", "Avg Time (ms)", "Steps", "Food Eaten", 
            "Deaths", "Win Rate (%)", "Score", "Level", "Timestamp"
        ]
        
        # Màu sắc cho styling Excel
        self.colors = {
            'header': '366092',      # Xanh đậm
            'algorithm': 'D9E1F2',  # Xanh nhạt
            'data': 'FFFFFF',       # Trắng
            'border': '000000'      # Đen
        }
    
    def start_game_session(self, algorithm_name: str):
        """
        Bắt đầu một session game mới
        Args:
            algorithm_name: Tên thuật toán (BFS, DFS, A*, UCS, IDS, Greedy)
        """
        self.current_game_data = {
            'algorithm': algorithm_name,
            'start_time': time.time(),
            'steps': 0,
            'food_eaten': 0,
            'deaths': 0,
            'score': 0,
            'level': 1,
            'is_win': False,
            'path_length': 0
        }
        print(f"Started game session with {algorithm_name}")
    
    def update_game_stats(self, **kwargs):
        """
        Cập nhật thống kê game hiện tại
        Args:
            **kwargs: Các thông số cần cập nhật (steps, food_eaten, deaths, score, level, etc.)
        """
        for key, value in kwargs.items():
            if key in self.current_game_data:
                if key == 'steps':
                    self.current_game_data['steps'] += value
                elif key == 'food_eaten':
                    self.current_game_data['food_eaten'] += value
                elif key == 'deaths':
                    self.current_game_data['deaths'] += value
                else:
                    self.current_game_data[key] = value
                print(f"Updated {key}: {self.current_game_data[key]}")
    
    def end_game_session(self, is_win: bool = False, final_score: int = 0):
        """
        Kết thúc session game hiện tại và lưu dữ liệu
        Args:
            is_win: Game có thắng không
            final_score: Điểm số cuối cùng
        """
        if not self.current_game_data:
            print("No active game session to end")
            return
        
        # Tính toán thời gian chạy
        end_time = time.time()
        avg_time_ms = (end_time - self.current_game_data['start_time']) * 1000
        
        # Tạo bản ghi hoàn chỉnh
        record = {
            'algorithm': self.current_game_data['algorithm'],
            'avg_time_ms': round(avg_time_ms, 2),
            'steps': self.current_game_data['steps'],
            'food_eaten': self.current_game_data['food_eaten'],
            'deaths': self.current_game_data['deaths'],
            'win_rate': 100.0 if is_win else 0.0,
            'score': final_score,
            'level': self.current_game_data['level'],
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Thêm vào danh sách
        self.data_records.append(record)
        
        # Reset current game data
        self.current_game_data = {}
        
        print(f"Game session ended - Algorithm: {record['algorithm']}, "
              f"Time: {record['avg_time_ms']}ms, Steps: {record['steps']}, "
              f"Food: {record['food_eaten']}, Deaths: {record['deaths']}, "
              f"Win: {'Yes' if is_win else 'No'}")
    
    def export_to_excel(self) -> bool:
        """
        Xuất dữ liệu ra file Excel
        Returns:
            True nếu thành công, False nếu có lỗi
        """
        try:
            # Kiểm tra file có tồn tại không
            file_exists = os.path.exists(self.excel_file)
            
            if file_exists:
                # Load workbook hiện có
                workbook = load_workbook(self.excel_file)
                worksheet = workbook.active
                print(f"Loading existing file: {self.excel_file}")
            else:
                # Tạo workbook mới
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Algorithm Comparison"
                
                # Thêm headers
                for col, header in enumerate(self.headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    self._style_header_cell(cell)
                
                print(f"Creating new file: {self.excel_file}")
            
            # Tìm dòng cuối cùng có dữ liệu
            last_row = worksheet.max_row
            if last_row == 1 and not worksheet.cell(row=1, column=1).value:
                # File trống, bắt đầu từ dòng 2
                start_row = 2
            else:
                # File có dữ liệu, thêm vào cuối
                start_row = last_row + 1
            
            # Thêm dữ liệu mới
            for i, record in enumerate(self.data_records):
                row = start_row + i
                worksheet.cell(row=row, column=1, value=record['algorithm'])
                worksheet.cell(row=row, column=2, value=record['avg_time_ms'])
                worksheet.cell(row=row, column=3, value=record['steps'])
                worksheet.cell(row=row, column=4, value=record['food_eaten'])
                worksheet.cell(row=row, column=5, value=record['deaths'])
                worksheet.cell(row=row, column=6, value=record['win_rate'])
                worksheet.cell(row=row, column=7, value=record['score'])
                worksheet.cell(row=row, column=8, value=record['level'])
                worksheet.cell(row=row, column=9, value=record['timestamp'])
                
                # Style cho dòng dữ liệu
                self._style_data_row(worksheet, row)
            
            # Auto-fit columns
            self._auto_fit_columns(worksheet)
            
            # Lưu file
            workbook.save(self.excel_file)
            
            print(f"Exported {len(self.data_records)} records to {self.excel_file}")
            print(f"Total records in file: {worksheet.max_row - 1}")
            
            # Không xóa data_records để có thể export nhiều lần
            # self.data_records.clear()  # Comment out để giữ dữ liệu
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def _style_header_cell(self, cell):
        """Style cho header cell"""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
    
    def _style_data_row(self, worksheet, row):
        """Style cho dòng dữ liệu"""
        for col in range(1, len(self.headers) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color=self.colors['border']),
                right=Side(style='thin', color=self.colors['border']),
                top=Side(style='thin', color=self.colors['border']),
                bottom=Side(style='thin', color=self.colors['border'])
            )
            
            # Alternating row colors
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type="solid")
            else:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
    
    def _auto_fit_columns(self, worksheet):
        """Tự động điều chỉnh độ rộng cột"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)  # Max width 20
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_statistics_summary(self) -> Dict:
        """
        Lấy thống kê tổng hợp của tất cả algorithms
        Returns:
            Dictionary chứa thống kê tổng hợp
        """
        if not self.data_records:
            return {}
        
        # Nhóm theo algorithm
        algorithm_stats = {}
        for record in self.data_records:
            algo = record['algorithm']
            if algo not in algorithm_stats:
                algorithm_stats[algo] = {
                    'total_games': 0,
                    'total_wins': 0,
                    'total_time': 0,
                    'total_steps': 0,
                    'total_food': 0,
                    'total_deaths': 0,
                    'total_score': 0
                }
            
            stats = algorithm_stats[algo]
            stats['total_games'] += 1
            if record['win_rate'] > 0:
                stats['total_wins'] += 1
            stats['total_time'] += record['avg_time_ms']
            stats['total_steps'] += record['steps']
            stats['total_food'] += record['food_eaten']
            stats['total_deaths'] += record['deaths']
            stats['total_score'] += record['score']
        
        # Tính trung bình
        summary = {}
        for algo, stats in algorithm_stats.items():
            summary[algo] = {
                'games_played': stats['total_games'],
                'win_rate': round((stats['total_wins'] / stats['total_games']) * 100, 2),
                'avg_time': round(stats['total_time'] / stats['total_games'], 2),
                'avg_steps': round(stats['total_steps'] / stats['total_games'], 2),
                'avg_food': round(stats['total_food'] / stats['total_games'], 2),
                'avg_deaths': round(stats['total_deaths'] / stats['total_games'], 2),
                'avg_score': round(stats['total_score'] / stats['total_games'], 2)
            }
        
        return summary
    
    def print_summary(self):
        """In thống kê tổng hợp ra console"""
        summary = self.get_statistics_summary()
        if not summary:
            print("No data to summarize")
            return
        
        print("\n" + "="*80)
        print("ALGORITHM PERFORMANCE SUMMARY")
        print("="*80)
        
        for algo, stats in summary.items():
            print(f"\n{algo}:")
            print(f"   Games Played: {stats['games_played']}")
            print(f"   Win Rate: {stats['win_rate']}%")
            print(f"   Avg Time: {stats['avg_time']}ms")
            print(f"   Avg Steps: {stats['avg_steps']}")
            print(f"   Avg Food: {stats['avg_food']}")
            print(f"   Avg Deaths: {stats['avg_deaths']}")
            print(f"   Avg Score: {stats['avg_score']}")
        
        print("\n" + "="*80)
